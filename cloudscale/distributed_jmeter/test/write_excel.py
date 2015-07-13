import re
from openpyxl import Workbook
from openpyxl import load_workbook
import os

if __name__ == "__main__":
    excel_file = '/Volumes/Storage/Xlab/cloudscale/measurement-results/Cap-shifted-minute/CloudStore.Measurements.Spring.2015.xlsx'
    wb = load_workbook(excel_file)
    path = "/Volumes/Storage/Xlab/cloudscale/gitlab/distributed-jmeter-standalone/bin/results"
    for f in os.listdir(path):
        file_path = '%s/%s/graphs/report.txt' % (path, f)
        if not os.path.isfile(f) and os.path.exists(file_path):
            with open(file_path) as fp:
                db_cpu_line = fp.next()
                fr_cpu_line = fp.next()
                vus60_line = fp.next()
                vus10_line = fp.next()
                vus5_line = fp.next()

                db_cpu = re.search("DBU 1m ?= ?(.*?)\n", db_cpu_line).group(1)
                fr_cpu = re.search("FU 1m ?= ?(.*?)\n", fr_cpu_line).group(1)
                vus60 = re.search("VU 1m ?= ?(.*?)\n", vus60_line).group(1)
                vus10 = re.search("VU 10s ?= ?(.*?)\n", vus10_line).group(1)
                vus5 = re.search("VU 5s ?= ?(.*?)\n", vus5_line).group(1)

            # Choose "Shifted capacity" worksheet
            ws = wb.get_sheet_by_name('Shifted capacity')
            printted = False
            for i in range(1, 147):
                cell = ws.cell(coordinate='B%s' % i)
                if cell.value == f:
                    try:
                        db_cpu = float(db_cpu)
                    except ValueError:
                        pass
                    ws['F%s' % i].value = db_cpu
                    ws['G%s' % i].value = ""
                    ws['H%s' % i].value = ""
                    ws['N%s' % i].value = float(fr_cpu)
                    ws['O%s' % i].value = ""
                    ws['P%s' % i].value = ""
                    ws['S%s' % i].value = int(float(vus60))
                    ws['T%s' % i].value = int(float(vus10))
                    ws['U%s' % i].value = int(float(vus5))
                    printted = True
                    break
            if not printted:
                print f
    wb.save(excel_file)
